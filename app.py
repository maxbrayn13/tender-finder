from flask import Flask, jsonify, request, send_from_directory, send_file, session
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
from datetime import datetime

app = Flask(__name__, static_folder='static')
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
CORS(app, supports_credentials=True)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database paths
DATABASE_PATH = os.getenv('DATABASE_PATH', 'goszakup_lots.db')
USERS_DB_PATH = os.getenv('USERS_DB_PATH', 'users.db')

# User model
class User(UserMixin):
    def __init__(self, id, username, email, is_admin=False):
        self.id = id
        self.username = username
        self.email = email
        self.is_admin = is_admin

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect(USERS_DB_PATH)
    conn.row_factory = sqlite3.Row
    user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()
    if user:
        return User(user['id'], user['username'], user['email'], user['is_admin'])
    return None

def init_users_db():
    """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    conn = sqlite3.connect(USERS_DB_PATH)
    cursor = conn.cursor()
    
    # –¢–∞–±–ª–∏—Ü–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # –°–æ–∑–¥–∞—Ç—å –∞–¥–º–∏–Ω–∞ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é
    admin_exists = cursor.execute('SELECT id FROM users WHERE username = ?', ('admin',)).fetchone()
    if not admin_exists:
        admin_password = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO users (username, email, password, is_admin)
            VALUES (?, ?, ?, ?)
        ''', ('admin', 'admin@tenderfinder.kz', admin_password, 1))
    
    conn.commit()
    conn.close()

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ
init_users_db()

def init_features_db():
    """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è —Ç–∞–±–ª–∏—Ü –¥–ª—è –Ω–æ–≤—ã—Ö —Ñ—É–Ω–∫—Ü–∏–π"""
    conn = sqlite3.connect(USERS_DB_PATH)
    cursor = conn.cursor()
    
    # 1. –ò–∑–±—Ä–∞–Ω–Ω–æ–µ
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS favorites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            lot_id INTEGER NOT NULL,
            added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, lot_id)
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_favorites_user ON favorites(user_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_favorites_lot ON favorites(lot_id)')
    
    # 2. –ò—Å—Ç–æ—Ä–∏—è –ø—Ä–æ—Å–º–æ—Ç—Ä–æ–≤
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS view_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            lot_id INTEGER NOT NULL,
            viewed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_view_history_user ON view_history(user_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_view_history_lot ON view_history(lot_id)')
    
    # 3. –ó–∞–º–µ—Ç–∫–∏
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            lot_id INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, lot_id)
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_user_notes_user ON user_notes(user_id)')
    
    # 4. –í—ã–∏–≥—Ä–∞–Ω–Ω—ã–µ —Ç–µ–Ω–¥–µ—Ä—ã
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS won_tenders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            lot_id INTEGER NOT NULL,
            won_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            actual_profit REAL,
            expected_profit REAL,
            notes TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, lot_id)
        )
    ''')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_won_tenders_user ON won_tenders(user_id)')
    
    conn.commit()
    conn.close()

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –Ω–æ–≤—ã—Ö —Ç–∞–±–ª–∏—Ü
init_features_db()

def get_db():
    """–ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ –∫ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö –ª–æ—Ç–æ–≤"""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_users_db():
    """–ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ –∫ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    conn = sqlite3.connect(USERS_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def calculate_stats(lot):
    """–†–∞—Å—á—ë—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –¥–ª—è –ª–æ—Ç–∞"""
    best_price = lot['tender_price'] * 0.4
    total_cost = best_price * lot['quantity']
    delivery_cost = total_cost * 0.15
    total_expense = total_cost + delivery_cost
    revenue = lot['tender_price'] * lot['quantity']
    profit = revenue - total_expense
    margin_percent = ((lot['tender_price'] - best_price) / best_price) * 100
    roi = (profit / total_expense) * 100 if total_expense > 0 else 0
    
    return {
        "best_price": round(best_price, 2),
        "total_cost": round(total_cost, 2),
        "delivery_cost": round(delivery_cost, 2),
        "total_expense": round(total_expense, 2),
        "revenue": round(revenue, 2),
        "profit": round(profit, 2),
        "margin_percent": round(margin_percent, 2),
        "roi": round(roi, 2)
    }

# ============= FRONTEND ROUTES =============

@app.route('/')
def index():
    """–ì–ª–∞–≤–Ω–∞—è —Å—Ç—Ä–∞–Ω–∏—Ü–∞"""
    return send_from_directory('static', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    """Serve static files"""
    if os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory('static', path)
    return send_from_directory('static', 'index.html')

# ============= AUTH API ROUTES =============

@app.route('/api/auth/register', methods=['POST'])
def register():
    """–†–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—è –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    try:
        data = request.get_json()
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        
        if not username or not email or not password:
            return jsonify({"error": "–í—Å–µ –ø–æ–ª—è –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã"}), 400
        
        conn = get_users_db()
        
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞–Ω–∏—è
        existing = conn.execute('SELECT id FROM users WHERE username = ? OR email = ?', 
                              (username, email)).fetchone()
        if existing:
            conn.close()
            return jsonify({"error": "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç"}), 400
        
        # –°–æ–∑–¥–∞–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
        hashed_password = generate_password_hash(password)
        cursor = conn.execute('''
            INSERT INTO users (username, email, password, is_admin)
            VALUES (?, ?, ?, ?)
        ''', (username, email, hashed_password, 0))
        
        user_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–π –≤—Ö–æ–¥
        user = User(user_id, username, email, False)
        login_user(user)
        
        return jsonify({
            "message": "–†–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—è —É—Å–ø–µ—à–Ω–∞",
            "user": {
                "id": user_id,
                "username": username,
                "email": email,
                "is_admin": False
            }
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """–í—Ö–æ–¥ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({"error": "–í–≤–µ–¥–∏—Ç–µ –ª–æ–≥–∏–Ω –∏ –ø–∞—Ä–æ–ª—å"}), 400
        
        conn = get_users_db()
        user_data = conn.execute('SELECT * FROM users WHERE username = ?', 
                                (username,)).fetchone()
        conn.close()
        
        if not user_data or not check_password_hash(user_data['password'], password):
            return jsonify({"error": "–ù–µ–≤–µ—Ä–Ω—ã–π –ª–æ–≥–∏–Ω –∏–ª–∏ –ø–∞—Ä–æ–ª—å"}), 401
        
        user = User(user_data['id'], user_data['username'], 
                   user_data['email'], user_data['is_admin'])
        login_user(user)
        
        return jsonify({
            "message": "–í—Ö–æ–¥ –≤—ã–ø–æ–ª–Ω–µ–Ω",
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "is_admin": user.is_admin
            }
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
@login_required
def logout():
    """–í—ã—Ö–æ–¥ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    logout_user()
    return jsonify({"message": "–í—ã—Ö–æ–¥ –≤—ã–ø–æ–ª–Ω–µ–Ω"})

@app.route('/api/auth/me')
@login_required
def get_current_user():
    """–ü–æ–ª—É—á–∏—Ç—å —Ç–µ–∫—É—â–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    return jsonify({
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "is_admin": current_user.is_admin
    })

# ============= EC2 DATA RECEIVING ENDPOINT =============

@app.route('/api/products/batch', methods=['POST'])
def receive_products():
    """
    –ü—Ä–∏–Ω—è—Ç—å —Ç–æ–≤–∞—Ä—ã –æ—Ç EC2 wrapper —Å–∫—Ä–∏–ø—Ç–∞
    
    –§–æ—Ä–º–∞—Ç –∑–∞–ø—Ä–æ—Å–∞:
    {
        "lot_number": "123456-–ó–¶–ü1",
        "lot_info": {
            "original_name": "...",
            "simplified_name": "...",
            ...
        },
        "products": [
            {
                "title": "Product name",
                "url": "https://...",
                "price": "100",
                "marketplace": "aliexpress.com",
                "method": "search"
            },
            ...
        ]
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        lot_number = data.get('lot_number')
        products = data.get('products', [])
        lot_info = data.get('lot_info', {})
        
        if not lot_number:
            return jsonify({"error": "Missing lot_number"}), 400
        
        if not products:
            return jsonify({"error": "Missing products"}), 400
        
        # –õ–æ–≥–∏—Ä—É–µ–º –ø–æ–ª—É—á–µ–Ω–∏–µ
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] üì• –ü–æ–ª—É—á–µ–Ω–æ –æ—Ç EC2:")
        print(f"   –õ–æ—Ç: {lot_number}")
        print(f"   –¢–æ–≤–∞—Ä–æ–≤: {len(products)}")
        if lot_info.get('simplified_name'):
            print(f"   –ù–∞–∑–≤–∞–Ω–∏–µ: {lot_info['simplified_name'][:50]}...")
        
        # –ó–¥–µ—Å—å –º–æ–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –≤ –ë–î –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
        # –ù–∞–ø—Ä–∏–º–µ—Ä, –æ–±–Ω–æ–≤–∏—Ç—å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ª–æ—Ç–µ –∏–ª–∏ –¥–æ–±–∞–≤–∏—Ç—å —Ç–æ–≤–∞—Ä—ã
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º —É—Å–ø–µ—à–Ω—ã–π –æ—Ç–≤–µ—Ç
        return jsonify({
            "status": "success",
            "message": "Products received successfully",
            "lot_number": lot_number,
            "products_count": len(products),
            "timestamp": timestamp
        }), 200
        
    except Exception as e:
        error_msg = str(e)
        print(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏—ë–º–∞ –¥–∞–Ω–Ω—ã—Ö: {error_msg}")
        return jsonify({
            "status": "error",
            "error": error_msg
        }), 500

# ============= ADMIN API ROUTES =============

@app.route('/api/admin/users')
@login_required
def admin_get_users():
    """–ü–æ–ª—É—á–∏—Ç—å –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (—Ç–æ–ª—å–∫–æ –¥–ª—è –∞–¥–º–∏–Ω–∞)"""
    if not current_user.is_admin:
        return jsonify({"error": "–î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â—ë–Ω"}), 403
    
    try:
        conn = get_users_db()
        users = conn.execute('SELECT id, username, email, is_admin, created_at FROM users ORDER BY created_at DESC').fetchall()
        conn.close()
        
        return jsonify({
            "users": [dict(user) for user in users]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
@login_required
def admin_toggle_admin(user_id):
    """–ü–µ—Ä–µ–∫–ª—é—á–∏—Ç—å —Å—Ç–∞—Ç—É—Å –∞–¥–º–∏–Ω–∞ (—Ç–æ–ª—å–∫–æ –¥–ª—è –∞–¥–º–∏–Ω–∞)"""
    if not current_user.is_admin:
        return jsonify({"error": "–î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â—ë–Ω"}), 403
    
    if user_id == current_user.id:
        return jsonify({"error": "–ù–µ–ª—å–∑—è –∏–∑–º–µ–Ω–∏—Ç—å —Å–≤–æ–∏ –ø—Ä–∞–≤–∞"}), 400
    
    try:
        conn = get_users_db()
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        
        if not user:
            conn.close()
            return jsonify({"error": "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω"}), 404
        
        new_status = 0 if user['is_admin'] else 1
        conn.execute('UPDATE users SET is_admin = ? WHERE id = ?', (new_status, user_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            "message": "–°—Ç–∞—Ç—É—Å –æ–±–Ω–æ–≤–ª—ë–Ω",
            "is_admin": bool(new_status)
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@login_required
def admin_delete_user(user_id):
    """–£–¥–∞–ª–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (—Ç–æ–ª—å–∫–æ –¥–ª—è –∞–¥–º–∏–Ω–∞)"""
    if not current_user.is_admin:
        return jsonify({"error": "–î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â—ë–Ω"}), 403
    
    if user_id == current_user.id:
        return jsonify({"error": "–ù–µ–ª—å–∑—è —É–¥–∞–ª–∏—Ç—å —Å–µ–±—è"}), 400
    
    try:
        conn = get_users_db()
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({"message": "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–¥–∞–ª—ë–Ω"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ============= PUBLIC API ROUTES =============

@app.route('/api')
def api_root():
    """API –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è"""
    return jsonify({
        "message": "TenderFinder API v2.0 with Auth",
        "authenticated": current_user.is_authenticated,
        "endpoints": {
            "auth": {
                "register": "/api/auth/register",
                "login": "/api/auth/login",
                "logout": "/api/auth/logout",
                "me": "/api/auth/me"
            },
            "public": {
                "health": "/api/health",
                "stats": "/api/stats"
            },
            "protected": {
                "lots": "/api/lots",
                "search": "/api/lots/search-by-budget"
            },
            "admin": {
                "users": "/api/admin/users"
            }
        }
    })

@app.route('/api/health')
def health():
    """Health check"""
    return jsonify({
        "status": "healthy",
        "database": os.path.exists(DATABASE_PATH),
        "users_db": os.path.exists(USERS_DB_PATH)
    })

# ============= PROTECTED API ROUTES =============

@app.route('/api/lots')
@login_required
def get_lots():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –ª–æ—Ç–æ–≤ (—Ç—Ä–µ–±—É–µ—Ç –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏)"""
    try:
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))
        category = request.args.get('category', '')
        search = request.args.get('search', '')
        
        conn = get_db()
        
        query = "SELECT * FROM lots WHERE 1=1"
        params = []
        
        if category:
            query += " AND category = ?"
            params.append(category)
        
        if search:
            query += " AND (simplified_name LIKE ? OR chinese_name LIKE ? OR lot_number LIKE ?)"
            search_term = f"%{search}%"
            params.extend([search_term, search_term, search_term])
        
        query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        
        lots = conn.execute(query, params).fetchall()
        
        count_query = "SELECT COUNT(*) as total FROM lots WHERE 1=1"
        count_params = []
        if category:
            count_query += " AND category = ?"
            count_params.append(category)
        if search:
            count_query += " AND (simplified_name LIKE ? OR chinese_name LIKE ? OR lot_number LIKE ?)"
            search_term = f"%{search}%"
            count_params.extend([search_term, search_term, search_term])
        
        total = conn.execute(count_query, count_params).fetchone()['total']
        conn.close()
        
        return jsonify({
            "total": total,
            "limit": limit,
            "offset": offset,
            "results": [dict(lot) for lot in lots]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/lots/<int:lot_id>')
@login_required
def get_lot(lot_id):
    """–ü–æ–ª—É—á–∏—Ç—å –ª–æ—Ç –ø–æ ID (—Ç—Ä–µ–±—É–µ—Ç –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏)"""
    try:
        conn = get_db()
        lot = conn.execute("SELECT * FROM lots WHERE id = ?", (lot_id,)).fetchone()
        
        if not lot:
            conn.close()
            return jsonify({"error": "–õ–æ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω"}), 404
        
        lot_dict = dict(lot)
        lot_dict['stats'] = calculate_stats(lot_dict)
        
        # –ü–æ–ª—É—á–∏—Ç—å —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–∏—Å–∫–∞ (—Å—Å—ã–ª–∫–∏ –Ω–∞ –º–∞–≥–∞–∑–∏–Ω—ã)
        search_results = conn.execute(
            "SELECT * FROM search_results WHERE lot_id = ? ORDER BY product_price",
            (lot_id,)
        ).fetchall()
        lot_dict['search_results'] = [dict(sr) for sr in search_results]
        
        conn.close()
        
        return jsonify(lot_dict)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/lots/search-by-budget', methods=['POST'])
@login_required
def search_by_budget():
    """–ü–æ–∏—Å–∫ –ª–æ—Ç–æ–≤ –ø–æ –±—é–¥–∂–µ—Ç—É (—Ç—Ä–µ–±—É–µ—Ç –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏)"""
    try:
        data = request.get_json()
        budget = float(data.get('budget', 0))
        
        min_budget = budget * 0.8
        max_budget = budget * 1.2
        
        conn = get_db()
        lots = conn.execute("SELECT * FROM lots").fetchall()
        conn.close()
        
        results = []
        for lot in lots:
            lot_dict = dict(lot)
            stats = calculate_stats(lot_dict)
            if min_budget <= stats['total_expense'] <= max_budget:
                lot_dict['stats'] = stats
                results.append(lot_dict)
        
        results.sort(key=lambda x: x['stats']['roi'], reverse=True)
        
        return jsonify({
            "query": {
                "budget": budget,
                "range": [min_budget, max_budget]
            },
            "total": len(results),
            "results": results[:20]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/lots/search-by-margin', methods=['POST'])
@login_required
def search_by_margin():
    """–ü–æ–∏—Å–∫ –ª–æ—Ç–æ–≤ –ø–æ –º–∞—Ä–∂–µ (—Ç—Ä–µ–±—É–µ—Ç –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏)"""
    try:
        data = request.get_json()
        target_margin = float(data.get('target_margin', 0))
        
        min_margin = target_margin * 0.8
        max_margin = target_margin * 1.2
        
        conn = get_db()
        lots = conn.execute("SELECT * FROM lots").fetchall()
        conn.close()
        
        results = []
        for lot in lots:
            lot_dict = dict(lot)
            stats = calculate_stats(lot_dict)
            if min_margin <= stats['profit'] <= max_margin:
                lot_dict['stats'] = stats
                results.append(lot_dict)
        
        results.sort(key=lambda x: x['stats']['roi'], reverse=True)
        
        return jsonify({
            "query": {
                "target_margin": target_margin,
                "range": [min_margin, max_margin]
            },
            "total": len(results),
            "results": results[:20]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/stats')
def get_stats():
    """–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ (–ø—É–±–ª–∏—á–Ω—ã–π –¥–æ—Å—Ç—É–ø)"""
    try:
        conn = get_db()
        
        total = conn.execute("SELECT COUNT(*) as total FROM lots").fetchone()['total']
        total_sum = conn.execute("SELECT SUM(tender_price * quantity) as total_sum FROM lots").fetchone()['total_sum'] or 0
        
        lots = conn.execute("SELECT * FROM lots").fetchall()
        conn.close()
        
        margins = [calculate_stats(dict(lot))['margin_percent'] for lot in lots]
        avg_margin = sum(margins) / len(margins) if margins else 0
        hot_deals = len([m for m in margins if m > 100])
        
        return jsonify({
            "total_lots": total,
            "total_sum": round(total_sum, 2),
            "avg_margin": round(avg_margin, 2),
            "hot_deals": hot_deals
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/categories')
def get_categories():
    """–ö–∞—Ç–µ–≥–æ—Ä–∏–∏ (–ø—É–±–ª–∏—á–Ω—ã–π –¥–æ—Å—Ç—É–ø)"""
    try:
        conn = get_db()
        categories = conn.execute("SELECT DISTINCT category FROM lots WHERE category IS NOT NULL").fetchall()
        conn.close()
        
        return jsonify({
            "categories": [row['category'] for row in categories]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== –ò–ó–ë–†–ê–ù–ù–û–ï ====================

@app.route('/api/favorites', methods=['GET'])
@login_required
def get_favorites():
    """–ü–æ–ª—É—á–∏—Ç—å –≤—Å–µ –∏–∑–±—Ä–∞–Ω–Ω—ã–µ –ª–æ—Ç—ã"""
    try:
        conn_users = get_users_db()
        favorites = conn_users.execute(
            'SELECT lot_id, added_at FROM favorites WHERE user_id = ? ORDER BY added_at DESC',
            (current_user.id,)
        ).fetchall()
        conn_users.close()
        
        if not favorites:
            return jsonify({"results": [], "total": 0, "stats": {}})
        
        lot_ids = [f['lot_id'] for f in favorites]
        placeholders = ','.join('?' * len(lot_ids))
        
        conn_lots = get_db()
        lots = conn_lots.execute(f'SELECT * FROM lots WHERE id IN ({placeholders})', lot_ids).fetchall()
        
        # –ü–æ–ª—É—á–∏—Ç—å search_results –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –ª–æ—Ç–∞
        results = []
        total_expense = 0
        total_profit = 0
        
        for lot in lots:
            lot_dict = dict(lot)
            lot_dict['stats'] = calculate_stats(lot_dict)
            
            # –î–æ–±–∞–≤–∏—Ç—å search_results
            search_results = conn_lots.execute(
                'SELECT * FROM search_results WHERE lot_id = ? ORDER BY product_price',
                (lot['id'],)
            ).fetchall()
            lot_dict['search_results'] = [dict(sr) for sr in search_results]
            
            results.append(lot_dict)
            total_expense += lot_dict['stats']['total_expense']
            total_profit += lot_dict['stats']['profit']
        
        conn_lots.close()
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        avg_roi = sum([l['stats']['roi'] for l in results]) / len(results) if results else 0
        avg_margin = sum([l['stats']['margin_percent'] for l in results]) / len(results) if results else 0
        
        return jsonify({
            "results": results,
            "total": len(results),
            "stats": {
                "total_lots": len(results),
                "total_expense": round(total_expense, 2),
                "total_profit": round(total_profit, 2),
                "avg_roi": round(avg_roi, 2),
                "avg_margin": round(avg_margin, 2)
            }
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/favorites', methods=['POST'])
@login_required
def add_favorite():
    """–î–æ–±–∞–≤–∏—Ç—å –≤ –∏–∑–±—Ä–∞–Ω–Ω–æ–µ"""
    try:
        data = request.get_json()
        lot_id = data.get('lot_id')
        
        conn = get_users_db()
        cursor = conn.cursor()
        
        try:
            cursor.execute(
                'INSERT INTO favorites (user_id, lot_id) VALUES (?, ?)',
                (current_user.id, lot_id)
            )
            conn.commit()
            conn.close()
            return jsonify({"success": True, "message": "–î–æ–±–∞–≤–ª–µ–Ω–æ –≤ –∏–∑–±—Ä–∞–Ω–Ω–æ–µ"})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({"error": "–£–∂–µ –≤ –∏–∑–±—Ä–∞–Ω–Ω–æ–º"}), 400
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/favorites/<int:lot_id>', methods=['DELETE'])
@login_required
def remove_favorite(lot_id):
    """–£–¥–∞–ª–∏—Ç—å –∏–∑ –∏–∑–±—Ä–∞–Ω–Ω–æ–≥–æ"""
    try:
        conn = get_users_db()
        conn.execute(
            'DELETE FROM favorites WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "–£–¥–∞–ª–µ–Ω–æ –∏–∑ –∏–∑–±—Ä–∞–Ω–Ω–æ–≥–æ"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/favorites/check/<int:lot_id>', methods=['GET'])
@login_required
def check_favorite(lot_id):
    """–ü—Ä–æ–≤–µ—Ä–∏—Ç—å –≤ –∏–∑–±—Ä–∞–Ω–Ω–æ–º –ª–∏ –ª–æ—Ç"""
    try:
        conn = get_users_db()
        favorite = conn.execute(
            'SELECT id FROM favorites WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        ).fetchone()
        conn.close()
        
        return jsonify({"is_favorite": favorite is not None})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== –ò–°–¢–û–†–ò–Ø –ü–†–û–°–ú–û–¢–†–û–í ====================

@app.route('/api/view-history', methods=['POST'])
@login_required
def add_view_history():
    """–ó–∞–ø–∏—Å–∞—Ç—å –ø—Ä–æ—Å–º–æ—Ç—Ä"""
    try:
        data = request.get_json()
        lot_id = data.get('lot_id')
        
        conn = get_users_db()
        conn.execute(
            'INSERT INTO view_history (user_id, lot_id) VALUES (?, ?)',
            (current_user.id, lot_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({"success": True})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/view-history/check', methods=['GET'])
@login_required
def get_viewed_lots():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –ø—Ä–æ—Å–º–æ—Ç—Ä–µ–Ω–Ω—ã—Ö –ª–æ—Ç–æ–≤"""
    try:
        conn = get_users_db()
        viewed = conn.execute(
            '''SELECT DISTINCT lot_id FROM view_history 
               WHERE user_id = ?''',
            (current_user.id,)
        ).fetchall()
        conn.close()
        
        return jsonify({
            "viewed_lot_ids": [v['lot_id'] for v in viewed]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== –ó–ê–ú–ï–¢–ö–ò ====================

@app.route('/api/notes', methods=['POST'])
@login_required
def save_note():
    """–°–æ–∑–¥–∞—Ç—å –∏–ª–∏ –æ–±–Ω–æ–≤–∏—Ç—å –∑–∞–º–µ—Ç–∫—É"""
    try:
        data = request.get_json()
        lot_id = data.get('lot_id')
        note = data.get('note', '').strip()
        
        if not note:
            return jsonify({"error": "–ó–∞–º–µ—Ç–∫–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç–æ–π"}), 400
        
        conn = get_users_db()
        cursor = conn.cursor()
        
        # –ü—Ä–æ–≤–µ—Ä–∏—Ç—å —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ –∑–∞–º–µ—Ç–∫–∞
        existing = cursor.execute(
            'SELECT id FROM user_notes WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        ).fetchone()
        
        if existing:
            # –û–±–Ω–æ–≤–∏—Ç—å
            cursor.execute(
                'UPDATE user_notes SET note = ?, updated_at = CURRENT_TIMESTAMP WHERE user_id = ? AND lot_id = ?',
                (note, current_user.id, lot_id)
            )
        else:
            # –°–æ–∑–¥–∞—Ç—å
            cursor.execute(
                'INSERT INTO user_notes (user_id, lot_id, note) VALUES (?, ?, ?)',
                (current_user.id, lot_id, note)
            )
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "–ó–∞–º–µ—Ç–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∞"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/notes/<int:lot_id>', methods=['GET'])
@login_required
def get_note(lot_id):
    """–ü–æ–ª—É—á–∏—Ç—å –∑–∞–º–µ—Ç–∫—É"""
    try:
        conn = get_users_db()
        note = conn.execute(
            'SELECT * FROM user_notes WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        ).fetchone()
        conn.close()
        
        if note:
            return jsonify(dict(note))
        return jsonify({"note": None})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/notes/<int:lot_id>', methods=['DELETE'])
@login_required
def delete_note(lot_id):
    """–£–¥–∞–ª–∏—Ç—å –∑–∞–º–µ—Ç–∫—É"""
    try:
        conn = get_users_db()
        conn.execute(
            'DELETE FROM user_notes WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "–ó–∞–º–µ—Ç–∫–∞ —É–¥–∞–ª–µ–Ω–∞"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/notes/check', methods=['GET'])
@login_required
def get_lots_with_notes():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –ª–æ—Ç–æ–≤ —Å –∑–∞–º–µ—Ç–∫–∞–º–∏"""
    try:
        conn = get_users_db()
        notes = conn.execute(
            'SELECT lot_id FROM user_notes WHERE user_id = ?',
            (current_user.id,)
        ).fetchall()
        conn.close()
        
        return jsonify({
            "lot_ids_with_notes": [n['lot_id'] for n in notes]
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== –í–´–ò–ì–†–ê–ù–ù–´–ï –¢–ï–ù–î–ï–†–´ ====================

@app.route('/api/won-tenders', methods=['POST'])
@login_required
def mark_as_won():
    """–û—Ç–º–µ—Ç–∏—Ç—å —Ç–µ–Ω–¥–µ—Ä –∫–∞–∫ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã–π"""
    try:
        data = request.get_json()
        lot_id = data.get('lot_id')
        actual_profit = data.get('actual_profit')
        expected_profit = data.get('expected_profit')
        notes = data.get('notes', '')
        
        conn = get_users_db()
        cursor = conn.cursor()
        
        try:
            cursor.execute(
                '''INSERT INTO won_tenders (user_id, lot_id, actual_profit, expected_profit, notes) 
                   VALUES (?, ?, ?, ?, ?)''',
                (current_user.id, lot_id, actual_profit, expected_profit, notes)
            )
            conn.commit()
            conn.close()
            return jsonify({"success": True, "message": "–¢–µ–Ω–¥–µ—Ä –æ—Ç–º–µ—á–µ–Ω –∫–∞–∫ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã–π"})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({"error": "–¢–µ–Ω–¥–µ—Ä —É–∂–µ –æ—Ç–º–µ—á–µ–Ω –∫–∞–∫ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã–π"}), 400
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/won-tenders/<int:lot_id>', methods=['PUT'])
@login_required
def update_won_tender(lot_id):
    """–û–±–Ω–æ–≤–∏—Ç—å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –≤—ã–∏–≥—Ä–∞–Ω–Ω–æ–º —Ç–µ–Ω–¥–µ—Ä–µ"""
    try:
        data = request.get_json()
        actual_profit = data.get('actual_profit')
        notes = data.get('notes')
        
        conn = get_users_db()
        conn.execute(
            'UPDATE won_tenders SET actual_profit = ?, notes = ? WHERE user_id = ? AND lot_id = ?',
            (actual_profit, notes, current_user.id, lot_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "–û–±–Ω–æ–≤–ª–µ–Ω–æ"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/won-tenders/<int:lot_id>', methods=['DELETE'])
@login_required
def delete_won_tender(lot_id):
    """–£–¥–∞–ª–∏—Ç—å –∏–∑ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã—Ö"""
    try:
        conn = get_users_db()
        conn.execute(
            'DELETE FROM won_tenders WHERE user_id = ? AND lot_id = ?',
            (current_user.id, lot_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "–£–¥–∞–ª–µ–Ω–æ –∏–∑ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã—Ö"})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/won-tenders/check', methods=['GET'])
@login_required
def get_won_lots():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –≤—ã–∏–≥—Ä–∞–Ω–Ω—ã—Ö –ª–æ—Ç–æ–≤"""
    try:
        conn = get_users_db()
        won = conn.execute(
            'SELECT lot_id, actual_profit FROM won_tenders WHERE user_id = ?',
            (current_user.id,)
        ).fetchall()
        conn.close()
        
        return jsonify({
            "won_lot_ids": [w['lot_id'] for w in won],
            "won_lots": {w['lot_id']: w['actual_profit'] for w in won}
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==================== –°–¢–ê–¢–ò–°–¢–ò–ö–ê ====================

@app.route('/api/user-stats', methods=['GET'])
@login_required
def get_user_stats():
    """–õ–∏—á–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    try:
        conn = get_users_db()
        
        # –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –∏–∑–±—Ä–∞–Ω–Ω—ã—Ö
        favorites_count = conn.execute(
            'SELECT COUNT(*) as count FROM favorites WHERE user_id = ?',
            (current_user.id,)
        ).fetchone()['count']
        
        # –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø—Ä–æ—Å–º–æ—Ç—Ä–µ–Ω–Ω—ã—Ö
        viewed_count = conn.execute(
            'SELECT COUNT(DISTINCT lot_id) as count FROM view_history WHERE user_id = ?',
            (current_user.id,)
        ).fetchone()['count']
        
        # –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–º–µ—Ç–æ–∫
        notes_count = conn.execute(
            'SELECT COUNT(*) as count FROM user_notes WHERE user_id = ?',
            (current_user.id,)
        ).fetchone()['count']
        
        # –í—ã–∏–≥—Ä–∞–Ω–Ω—ã–µ —Ç–µ–Ω–¥–µ—Ä—ã
        won_tenders = conn.execute(
            'SELECT * FROM won_tenders WHERE user_id = ?',
            (current_user.id,)
        ).fetchall()
        
        conn.close()
        
        won_count = len(won_tenders)
        total_actual_profit = sum([w['actual_profit'] or 0 for w in won_tenders])
        total_expected_profit = sum([w['expected_profit'] or 0 for w in won_tenders])
        avg_roi = ((total_actual_profit / total_expected_profit * 100) - 100) if total_expected_profit > 0 else 0
        
        return jsonify({
            "viewed_lots": viewed_count,
            "favorites_count": favorites_count,
            "notes_count": notes_count,
            "won_tenders": won_count,
            "total_actual_profit": round(total_actual_profit, 2),
            "total_expected_profit": round(total_expected_profit, 2),
            "avg_roi": round(avg_roi, 2),
            "win_rate": round((won_count / viewed_count * 100), 2) if viewed_count > 0 else 0
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ==================== EXCEL –≠–ö–°–ü–û–†–¢ ====================

@app.route('/api/favorites/export', methods=['GET'])
@login_required
def export_favorites_excel():
    """–≠–∫—Å–ø–æ—Ä—Ç –∏–∑–±—Ä–∞–Ω–Ω—ã—Ö –≤ Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        # –ü–æ–ª—É—á–∏—Ç—å –∏–∑–±—Ä–∞–Ω–Ω—ã–µ –ª–æ—Ç—ã
        conn_users = get_users_db()
        favorites = conn_users.execute(
            'SELECT lot_id FROM favorites WHERE user_id = ? ORDER BY added_at DESC',
            (current_user.id,)
        ).fetchall()
        conn_users.close()
        
        if not favorites:
            return jsonify({"error": "–ù–µ—Ç –∏–∑–±—Ä–∞–Ω–Ω—ã—Ö –ª–æ—Ç–æ–≤"}), 400
        
        lot_ids = [f['lot_id'] for f in favorites]
        placeholders = ','.join('?' * len(lot_ids))
        
        conn_lots = get_db()
        lots = conn_lots.execute(f'SELECT * FROM lots WHERE id IN ({placeholders})', lot_ids).fetchall()
        
        # –°–æ–∑–¥–∞—Ç—å Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "–ò–∑–±—Ä–∞–Ω–Ω—ã–µ –ª–æ—Ç—ã"
        
        # –ó–∞–≥–æ–ª–æ–≤–∫–∏
        headers = ['‚Ññ', '–ù–∞–∑–≤–∞–Ω–∏–µ', '–ö–∞—Ç–µ–≥–æ—Ä–∏—è', '–¢–µ–Ω–¥–µ—Ä (‚Ç∏)', '–ö–æ–ª-–≤–æ', '–ó–∞—Ç—Ä–∞—Ç—ã (‚Ç∏)', 
                   '–ü—Ä–∏–±—ã–ª—å (‚Ç∏)', 'ROI (%)', '–ú–∞—Ä–∂–∞ (%)', '–õ—É—á—à–∞—è —Ü–µ–Ω–∞ (‚Ç∏)']
        ws.append(headers)
        
        # –°—Ç–∏–ª—å –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # –î–∞–Ω–Ω—ã–µ
        total_expense = 0
        total_profit = 0
        
        for idx, lot in enumerate(lots, 1):
            lot_dict = dict(lot)
            stats = calculate_stats(lot_dict)
            
            row = [
                idx,
                lot_dict['simplified_name'] or lot_dict['original_name'],
                lot_dict['category'] or '–î—Ä—É–≥–æ–µ',
                round(lot_dict['tender_price'], 2),
                lot_dict['quantity'],
                round(stats['total_expense'], 2),
                round(stats['profit'], 2),
                round(stats['roi'], 2),
                round(stats['margin_percent'], 2),
                round(stats['best_price'], 2)
            ]
            ws.append(row)
            
            total_expense += stats['total_expense']
            total_profit += stats['profit']
        
        conn_lots.close()
        
        # –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç—Ä–æ–∫–∞
        avg_roi = (total_profit / total_expense * 100) if total_expense > 0 else 0
        avg_margin = sum([calculate_stats(dict(lot))['margin_percent'] for lot in lots]) / len(lots)
        
        total_row = [
            '',
            '–ò–¢–û–ì–û',
            '',
            '',
            '',
            round(total_expense, 2),
            round(total_profit, 2),
            round(avg_roi, 2),
            round(avg_margin, 2),
            ''
        ]
        ws.append(total_row)
        
        # –°—Ç–∏–ª—å –∏—Ç–æ–≥–æ–≤–æ–π —Å—Ç—Ä–æ–∫–∏
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # –°–æ—Ö—Ä–∞–Ω–∏—Ç—å –≤ BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'TenderFinder_Favorites_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
